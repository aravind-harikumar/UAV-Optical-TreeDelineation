"""
..   module:: Main Function
    :platform: Unix, Windows
    :synopsis: A module for preprocessing DEM
    
..  :author:: Aravind Harikumar <aravind.harikumar@utoronto.ca>
"""
import io, sys, os
# from Others import testRexecution
from openpyxl import load_workbook
from MetashapeProcessor import generateOrthophoto
from ProjectConstants import GlobalConstants as gc
# sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from ImageProcessor.RasterOperators import ReadImage
from ITCDetector import DataPreprocessor, \
        IndividualTreeDetection as ITC, \
        ITCAnalyze, \
        IndividualTreeDelineation as ITD, \
        IndividualTreeBuffering as ITB, \
        IndividualTreeSpanning as ITSpan, \
        IndividualTreeDelineationAll as ITD1
    
# Dates = ['20170808','20171011','20171013']
Dates = ['20180710'] # '20180514','20180710','20181015'
Base_RawData_Folder = '/media/ensmingerlabgpu/Seagate/2018/'
Base_Processed_Data_MS = '/mnt/4TBHDD/Spruce_Up_New_Copy/SpruceUp/UAV/StCasimir/2018/Nano-Hyperspec/ProcessedMSData/'
Base_Processed_Data_HS = '/mnt/4TBHDD/Spruce_Up_New_Copy/SpruceUp/UAV/StCasimir/2018/Nano-Hyperspec/ProcessedMSData/'

# Dates = ['20181015']
runpart = True
skipCoregStep = True

if(runpart):
    '''#################################################################'''
    '''           Generate Othorectfied Image and DEM                   '''
    '''#################################################################'''

#     wb = load_workbook(filename = 'StCasmir_Data.xlsx')
#     ws = wb['Sheet1']

#     for index, row in enumerate(ws.iter_rows()):
#         # skip header row
#         if(str(row[0].value).strip() in Dates):
#             print(row[0].value)
#         # if index != 5 and index != 5 and index != 5 :                                                                        
#         #     continue

#             # # set image details
#             DataInfo = {
#                 'ID'             : str(row[0].value),
#                 'DataPath'       : str(row[2].value),
#                 'ContolPointFile': '/mnt/4TBHDD/Spruce_Up_New_Copy/SpruceUp/GPS/only_targets/SCA_GPS_final_edit_petra.csv',
#                 'WorkingFolder'  : Base_RawData_Folder,
#                 }
#             DataObj = generateOrthophoto.MetaShapeProcessing(DataInfo)
#             # DataObj.GenerateOrthoImage()

#     '''#################################################################'''
#     '''                       Stack bands / Tile Image                  '''
#     '''#################################################################'''
#     # exit(0)
#     for date in Dates:
#         FileDataInfo = {
#             'RSDataType'        : 'Mulispectral', # Hyperspectral or Mulispectral
#             'BaseFolder'        :  Base_RawData_Folder,
#             'ImageFolder'       :  date, # 20180514, 20180710, 20181015
#             'OrthoPhoto'        :  date + '.tif',  # 20180514/20180514.tif , 20180710/20180710.tif , 20181015/20181015.tif
#             'nDSM'              : 'nDSM.tif', # '/mnt/4TBHDD/Agisoft/nDSM_Z19_UTM.tif',
#             'SelectedBands'     : ['Band1', 'Band2', 'Band3', 'Band4', 'Band5'], # or ALL
#             'OutFolderHyper'    :  Base_Processed_Data_HS,
#             'OutFolderMulti'    :  Base_Processed_Data_MS,
#             'RefSpanImage'      : '/mnt/4TBHDD/Agisoft/nDSM_Z19_UTM.tif',#'/home/ensmingerlabgpu/Desktop/AgisoftProjects/20180514/20180514.tif',
#             'RefnDSMImage'      : '/mnt/4TBHDD/Agisoft/20180514_ndsm.tif',
#             'StudyAreaShp'      : '/mnt/4TBHDD/Spruce_Up_New_Copy/SpruceUp/Site_info/ArcGIS/StCasimir_SpruceUp_temp.shp',
#             'CoBaseFolder'      :  Base_Processed_Data_MS,
#             'RefDateFolder'     : '20180514',
#             'CoregisterFolder'  :  gc.COREGISTER_OUT_FOLDER,
#             'SaveGCPFromAgisoft':  True,
#             'SaveAgisoftFile'   :  False
#         }
#         # Preprocess Data 
#         print('Processing - '  + FileDataInfo['ImageFolder'])
#         if(skipCoregStep):
#             FileDataInfo['CoregisterFolder']  = gc.NON_COR_ORTHOPHOTO_FOLDER

#         DataObj = DataPreprocessor.Preprocessor(FileDataInfo)
#         DataObj.ProcessData()

#     # Coregister Data From All Aquisitions
#     ITCDelinObj = ITD1.ITCDelineationUtils(FileDataInfo) # names of dsm and orthphoto should be same , only date  different in name
#     if(not(skipCoregStep)):
#         ITCDelinObj.Coregister()
#         ITCDelinObj.Coregister_orig_dim_dem()

#     #  Tile Orthophoto and nDSM Orig Res
#     for date in Dates:
#         if(skipCoregStep):
#             FileDataInfo['CoregisterFolder']  = gc.NON_COR_ORTHOPHOTO_FOLDER

#         FileDataInfo['ImageFolder'] = date
#         FileDataInfo['OrthoPhoto'] = date + '.tif'
#         print('Tiling -' + FileDataInfo['ImageFolder'])
#         print("here:" + FileDataInfo['CoregisterFolder'])
#         DataObj = DataPreprocessor.Preprocessor(FileDataInfo)
#         DataObj.TileData()

# '''###################################################################'''
# '''           Detect Individual Tree Crowns From Optical Data         '''
# '''###################################################################'''
# for date in Dates:
#     FileDataInfo = {
#         'BaseFolder'            :  Base_Processed_Data_MS,
#         'ImageDate'             :  date, # 20180514, 20180710, 20181015
#         'ITCBufferSize'         :  1.5,
#         'TreeTopInclusionBuffer':  2, # includes multiple bright points around N m of tree top
#         'InterTreeSepeartion'   :  1, # removes redundant points within Nm distance
#         'MinTreeHeight'         :  1, # (m)
#         'FilterVariance'        :  7
#     }
#     print('ITC Detecting -' + FileDataInfo['ImageDate'])
#     ITCObj = ITC.ITCUtils(FileDataInfo)
#     ITCObj.DetectIndividaulTrees()
# # exit(0)

#     '''#########################################################################################'''
#     '''             Get Spectral Maps                                                  '''
#     '''#########################################################################################'''
#     for date in Dates:
#         FileDataInfo = {
#             'BaseFolder'        :  Base_Processed_Data_MS,
#             'Date'              :  date, # 20180514, 20180710, 20181015
#             'ImageFolder'       :  gc.COREGISTER_OUT_FOLDER,
#             'InShapeFolder'     : 'ITC-Data/All-ITC',
#             'OutShapeFolder'    : 'ITC-Data/All-ITC/Out/',
#             'SpectralMaps'      : 'Maps',
#             'SkipCoregStep'     : skipCoregStep
#         }

#         if(skipCoregStep):
#             FileDataInfo['CoregisterFolder']  = gc.NON_COR_ORTHOPHOTO_FOLDER

#         print('ITC Delineating -' + FileDataInfo['Date'])
#         ITCObj = ITD.ITCDelineationUtils(FileDataInfo)
#         ITCObj.DelineateIndividaulTrees()

#         # exit(0)

    '''#########################################################################################'''
    '''                                 Get fuzzy maps of crown                                 '''
    '''#########################################################################################'''
    # Dates = ['20181015']
    for date in Dates:
        FileDataInfo = {
            'BaseFolder'        :  Base_Processed_Data_MS,
            'Date'              :  date, # 20180514, 20180710, 20181015
            'ClipSize'          :  4,
            # 'RefCenter'       :  [1117, 1398, 1024, 709, 9345], # old
            'RefCenter'         :  [1480, 1991, 1407, 919, 11280],
            'NumOfClusters'     :  2,
            'SkipCoregStep'     :  skipCoregStep
        }
        print('ITC Delineating -' + FileDataInfo['Date'])
        ITCObj = ITB.ITCBufferUtils(FileDataInfo)
        # ITCObj.BufferIndividaulTrees() # run Simple FCM
        # ITCObj.RBufferIndividaulTrees() # run MRF-FCM
    # exit(0)

    '''#########################################################################################'''
    '''                      Delineate Individual Tree Crowns From FCM Output                 '''
    '''#########################################################################################'''
    # Dates = ['20180514','20180710','20181015']
    for date in Dates:
        FileDataInfo = {
            'BaseFolder'        :  Base_Processed_Data_MS,
            'Date'              :  date,
            'RefShp'            : '/home/ensmingerlabgpu/Documents/MATLAB/BasicSnake_version2f/myfile_1.shp',
            'MatFilePath'       : '/home/ensmingerlabgpu/Documents/MATLAB/BasicSnake_version2f',
            'SkipCoregStep'     :  skipCoregStep,
            'IndexOperatorType' : 'MaxNPixelMean', # AllPixelMean, MaxNPixelMean
            'Skip'              :  False
        }
        print('ITC Delineating -' + FileDataInfo['Date'])
        ITCObj = ITSpan.ITCCrownShapeUtils(FileDataInfo)
        # ITCObj.GetCrownFromMatlab()
        # ITCObj.GetCrownBuffer()

# exit(0)

# '''#################################################################'''
# '''                          Analysis                               '''
# '''#################################################################'''

FileDataInfo = {
    'BaseFolder'        : Base_Processed_Data_MS,
    'Dates'             : ['20180514','20180710','20181015'],
    'MaxNeighbourDist'  : 2.5
}
print('Analyzing...')
AnayzeObj = ITCAnalyze.AnalyzeUtils(FileDataInfo)
AnayzeObj.ExtractPhenology()